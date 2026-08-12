#!/usr/bin/env python3
"""
Pima County Treasurer delinquency file -> Tax Delinquent document leads.

Feed-ready translator. The Treasurer sells a $50/month "all delinquent parcels"
data file (updated monthly); the annual tax-lien-sale list is also published.
This script turns whichever file is present into doc-type lead records, and
cleanly no-ops when no file is present (so the daily pipeline never fails just
because the subscription file hasn't been dropped in yet).

Looked-for inputs (first match wins), under data/:
    pima_tax_delinquent.xlsx   (the monthly subscription file)
    pima_tax_delinquent.csv
    pima_tax_delinquent.tsv

Column mapping is fuzzy (headers vary between the xlsx feed and the sale list):
    APN / parcel      -> apn
    amount / balance  -> amount_owed
    tax year          -> tax_year   (optional)
    owner / taxpayer  -> owner       (optional; parcel master fills this anyway)

Output (doc-type store, same shape the pipeline consumes):
    data/pima_tax_docs.jsonl        one "Tax Delinquent" doc per delinquent parcel

Usage:
    python pipeline/enrich_treasurer.py
    python pipeline/enrich_treasurer.py --input path/to/file.xlsx
    python pipeline/enrich_treasurer.py --as-of 2026-08-01   # recorded_date stamp
"""
from __future__ import annotations
import argparse
import csv
import json
import logging
import re
from datetime import date, datetime, timezone
from pathlib import Path

DATA_DIR = Path(__file__).resolve().parent.parent / "data"
STORE_PATH = DATA_DIR / "pima_tax_docs.jsonl"
DEFAULT_INPUTS = [
    DATA_DIR / "pima_tax_delinquent.xlsx",
    DATA_DIR / "pima_tax_delinquent.csv",
    DATA_DIR / "pima_tax_delinquent.tsv",
]

logging.basicConfig(level=logging.INFO,
                    format="%(asctime)s [%(levelname)s] %(message)s",
                    datefmt="%H:%M:%S")
log = logging.getLogger("enrich_treasurer")

APN_HINTS = ["apn", "parcel", "parcelnumber", "parcel number", "parcel_id",
             "parcelid", "taxcode", "tax code", "state code", "statecode"]
AMOUNT_HINTS = ["amount owed", "amount", "total due", "balance", "tax due",
                "delinquent", "delinquent tax", "owed", "total", "amount_owed"]
YEAR_HINTS = ["tax year", "taxyear", "year", "tax_year"]
OWNER_HINTS = ["owner", "taxpayer", "owner name", "name"]


def _norm(h) -> str:
    return re.sub(r"\s+", " ", str(h or "").lower().replace("_", " ").replace("-", " ")).strip()


def _pick(headers: list[str], hints: list[str]) -> str | None:
    norm = {_norm(h): h for h in headers}
    for hint in hints:                       # exact first
        if hint in norm:
            return norm[hint]
    for hint in hints:                       # then substring
        for n, orig in norm.items():
            if hint in n:
                return orig
    return None


def _rows_from_csv(path: Path, delim: str) -> tuple[list[str], list[dict]]:
    with path.open(newline="", encoding="utf-8-sig") as f:
        r = csv.DictReader(f, delimiter=delim)
        rows = list(r)
        return (r.fieldnames or []), rows


def _rows_from_xlsx(path: Path) -> tuple[list[str], list[dict]]:
    try:
        from openpyxl import load_workbook
    except ImportError:
        raise SystemExit("openpyxl required for xlsx input: pip install openpyxl")
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    it = ws.iter_rows(values_only=True)
    headers = [str(h) if h is not None else "" for h in next(it)]
    rows = [dict(zip(headers, vals)) for vals in it]
    return headers, rows


def parse_amount(v) -> float:
    if v is None:
        return 0.0
    n = re.sub(r"[^0-9.\-]", "", str(v))
    try:
        return float(n) if n else 0.0
    except ValueError:
        return 0.0


def norm_apn(v) -> str:
    return re.sub(r"[^A-Z0-9]", "", str(v or "").upper())


def run(input_path: Path | None, as_of: str) -> None:
    path = input_path
    if path is None:
        path = next((p for p in DEFAULT_INPUTS if p.exists()), None)

    if path is None or not path.exists():
        log.info("no Pima treasurer file present (looked for "
                 f"{', '.join(p.name for p in DEFAULT_INPUTS)}) — skipping, "
                 "Tax Delinquent leads will simply be absent this run")
        # feed-ready: leave any existing store untouched, exit clean
        return

    log.info(f"reading {path.name}")
    if path.suffix.lower() in (".xlsx", ".xlsm"):
        headers, rows = _rows_from_xlsx(path)
    else:
        delim = "\t" if path.suffix.lower() == ".tsv" else ","
        headers, rows = _rows_from_csv(path, delim)

    apn_col = _pick(headers, APN_HINTS)
    amt_col = _pick(headers, AMOUNT_HINTS)
    yr_col = _pick(headers, YEAR_HINTS)
    own_col = _pick(headers, OWNER_HINTS)
    if not apn_col:
        raise SystemExit(f"could not find an APN/parcel column in headers: {headers}")
    log.info(f"columns → apn={apn_col!r} amount={amt_col!r} year={yr_col!r} owner={own_col!r}")

    stamp = datetime.now(timezone.utc).isoformat(timespec="seconds").replace("+00:00", "Z")
    out, seen = [], set()
    for row in rows:
        apn_raw = row.get(apn_col)
        apn = norm_apn(apn_raw)
        if not apn or apn in seen:
            continue
        seen.add(apn)
        year = str(row.get(yr_col)).strip() if yr_col and row.get(yr_col) else None
        out.append({
            "county":        "Pima",
            "source":        "pima_treasurer",
            "doc_type":      "Tax Delinquent",
            "doc_code":      "TAXDLQ",
            "category":      "Tax & Liens",
            "doc_number":    f"TAX-{apn}" + (f"-{year}" if year else ""),
            "recorded_date": as_of,
            "names":         [str(row.get(own_col)).strip()] if own_col and row.get(own_col) else [],
            "apn":           str(apn_raw).strip() or None,
            "apn_norm":      apn,
            "resolved":      True,   # keyed by APN; parcel master fills address/owner
            "amount_owed":   parse_amount(row.get(amt_col)) if amt_col else None,
            "tax_year":      year,
            "fetched_at":    stamp,
        })

    STORE_PATH.parent.mkdir(parents=True, exist_ok=True)
    with STORE_PATH.open("w") as f:
        for r in out:
            f.write(json.dumps(r) + "\n")
    log.info(f"✓ wrote {len(out):,} Tax Delinquent docs → {STORE_PATH}")


def main(argv=None):
    ap = argparse.ArgumentParser(description=__doc__,
                                 formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("--input", type=Path, default=None,
                    help="Explicit path to the delinquency file (else auto-detect in data/)")
    ap.add_argument("--as-of", type=str, default=date.today().isoformat(),
                    help="recorded_date stamp for the emitted docs (default: today)")
    args = ap.parse_args(argv)
    run(args.input, args.as_of)


if __name__ == "__main__":
    main()
