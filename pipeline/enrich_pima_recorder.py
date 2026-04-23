#!/usr/bin/env python3
"""
enrich_pima_recorder.py
=======================

Translates the Pima County Recorder scraper output (CSV or XLSX) into
canonical signals JSONL that build_leads.py picks up automatically.

Input: data/pima_recorder.csv OR data/pima_recorder.xlsx
       (client's scraper output — columns include Document Type,
        Sequence Number, Recording Date, Grantors, Grantees, Parcel ID,
        Address, Legal Description)

Output: data/pima_recorder_raw.jsonl
        (same format as data/tax_delinquent_pima.jsonl — flows into
         the existing pipeline with no build_leads.py changes)

Usage:
    python pipeline/enrich_pima_recorder.py
    python pipeline/enrich_pima_recorder.py --input data/pima_recorder.xlsx
    python pipeline/enrich_pima_recorder.py --dry-run       # preview only
"""
from __future__ import annotations
import argparse
import csv
import json
import os
import re
import sys
from pathlib import Path
from typing import Optional, Dict, Iterable

ROOT = Path(__file__).resolve().parents[1]
DEFAULT_INPUTS = [
    ROOT / "data" / "pima_recorder.xlsx",
    ROOT / "data" / "pima_recorder.csv",
    ROOT / "data" / "Pimacounty_Data.csv",  # default name from client's script
]
OUTPUT_JSONL = ROOT / "data" / "pima_recorder_raw.jsonl"

# ---------------------------------------------------------------------------
# DOCUMENT TYPE → CANONICAL SIGNAL MAPPING
# ---------------------------------------------------------------------------
# Pima Recorder doc types vary in wording. This maps fuzzy substrings to the
# canonical signal names already weighted in pipeline/build_leads.py.
#
# Order matters — first match wins. Put more specific patterns first.
# ---------------------------------------------------------------------------

SIGNAL_MAPPINGS = [
    # Notice of Trustee Sale — pre-foreclosure
    (r"notice\s*of\s*trustee", "Notice of Trustee Sale"),
    (r"notice\s*of\s*sale",    "Notice of Trustee Sale"),
    (r"\bnots\b",              "Notice of Trustee Sale"),

    # Sheriff's Deed — post-foreclosure (property already sold at auction)
    (r"sheriff.*deed",         "Sheriffs Deed"),
    (r"trustee.*deed",         "Sheriffs Deed"),

    # Lis Pendens — lawsuit pending against the property
    (r"lis\s*pendens",         "Lis Pendens"),
    (r"notice\s*of\s*pending", "Lis Pendens"),

    # Judgment Lien — court judgment recorded against owner
    (r"judgment\s*lien",       "Judgment Lien"),
    (r"\bjudgment\b",          "Judgment Lien"),
    (r"\bjudgement\b",         "Judgment Lien"),

    # Medicaid Lien — AHCCCS is Arizona's Medicaid program
    (r"ahcccs",                "Medicaid Lien"),
    (r"medicaid",              "Medicaid Lien"),

    # Federal Tax Lien — IRS
    (r"federal\s*tax\s*lien",  "Federal Tax Lien"),
    (r"\birs\b.*lien",         "Federal Tax Lien"),

    # State Tax Lien — Arizona DOR
    (r"state\s*tax\s*lien",    "State Tax Lien"),
    (r"arizona.*tax.*lien",    "State Tax Lien"),

    # City Lien — municipal code / abatement / utility
    (r"city\s*lien",           "City Lien"),
    (r"municipal\s*lien",      "City Lien"),
    (r"code\s*enforcement",    "City Lien"),
    (r"abatement",             "City Lien"),

    # Hospital Lien
    (r"hospital\s*lien",       "Hospital Lien"),

    # Mechanics Lien — contractor unpaid
    (r"mechanic.*lien",        "Mechanics Lien"),

    # HOA Lien
    (r"hoa\s*lien",            "HOA Lien"),
    (r"homeowner.*assoc.*lien","HOA Lien"),

    # Bankruptcy
    (r"bankruptcy",            "Bankruptcy"),

    # Estate / Probate
    (r"affidavit.*death",      "Affidavit of Death"),
    (r"affidavit.*successor",  "Affidavit of Death"),
    (r"probate",               "Probate Case"),

    # Divorce
    (r"divorce",               "Divorce"),
    (r"dissolution.*marriage", "Divorce"),
]


def map_doc_type(raw: str) -> Optional[str]:
    """Map a raw Pima Recorder document type string to our canonical signal."""
    if not raw:
        return None
    text = raw.strip().lower()
    for pattern, signal in SIGNAL_MAPPINGS:
        if re.search(pattern, text):
            return signal
    return None


# ---------------------------------------------------------------------------
# APN NORMALIZATION (Pima: strip hyphens, uppercase)
# ---------------------------------------------------------------------------

def norm_apn(raw: str) -> Optional[str]:
    """Normalize a Pima APN. Pima uses formats like 219-21-627X or 21921627X."""
    if not raw:
        return None
    s = str(raw).strip().upper()
    # Strip common non-APN characters
    s = re.sub(r"[^\w]", "", s)
    if not s:
        return None
    # Pima APNs are typically 8-9 alphanumeric chars
    if len(s) < 6 or len(s) > 15:
        return None
    return s


# ---------------------------------------------------------------------------
# ADDRESS FALLBACK — match by site address if APN missing
# ---------------------------------------------------------------------------

# Common street-type abbreviations to normalize away
_STREET_SUFFIXES = {
    "STREET": "ST", "AVENUE": "AVE", "ROAD": "RD", "DRIVE": "DR",
    "BOULEVARD": "BLVD", "COURT": "CT", "CIRCLE": "CIR", "LANE": "LN",
    "PLACE": "PL", "PARKWAY": "PKWY", "HIGHWAY": "HWY", "TERRACE": "TER",
    "WAY": "WAY", "TRAIL": "TRL",
}
_DIRECTION_EXPAND = {"N": "N", "S": "S", "E": "E", "W": "W",
                     "NORTH": "N", "SOUTH": "S", "EAST": "E", "WEST": "W",
                     "NE": "NE", "NW": "NW", "SE": "SE", "SW": "SW"}


def clean_address(raw: str) -> Optional[str]:
    """Normalize an address for fuzzy matching.
    '500 W. Speedway Blvd, Tucson, AZ 85705' → '500 W SPEEDWAY BLVD'
    '1234 North Main Street Tucson AZ 85701' → '1234 N MAIN ST'
    """
    if not raw:
        return None
    s = str(raw).strip().upper()
    # Strip punctuation
    s = re.sub(r"[.,#]", " ", s)
    # Strip trailing zip (5 digits or 5+4)
    s = re.sub(r"\s+\d{5}(-\d{4})?\s*$", "", s)
    # Strip trailing state (2-letter code)
    s = re.sub(r"\s+[A-Z]{2}\s*$", "", s)
    # Strip common Arizona city names at tail
    for city in ("TUCSON", "PHOENIX", "MESA", "CHANDLER", "SCOTTSDALE",
                 "GLENDALE", "GILBERT", "TEMPE", "PEORIA", "SURPRISE",
                 "AVONDALE", "GOODYEAR", "BUCKEYE", "ORO VALLEY", "MARANA",
                 "SAHUARITA", "VAIL", "GREEN VALLEY", "CATALINA"):
        s = re.sub(rf"\s+{city}\s*$", "", s)
    # Strip unit/apt/suite
    s = re.sub(r"\s+(APT|UNIT|SUITE|STE|LOT)\s+\S+\s*$", "", s)
    # Collapse whitespace
    s = re.sub(r"\s+", " ", s).strip()
    if not s:
        return None
    tokens = s.split()
    # Normalize directions (any position)
    tokens = [_DIRECTION_EXPAND.get(t, t) for t in tokens]
    # Normalize street-type suffixes (any position, but most commonly last)
    tokens = [_STREET_SUFFIXES.get(t, t) for t in tokens]
    return " ".join(tokens)


def load_address_index(parcels_path: Path) -> Dict[str, str]:
    """Build a {cleaned_address: apn} index from pima_parcels.jsonl.
    Used for fallback when a recorder row has no parcel ID."""
    index: Dict[str, str] = {}
    if not parcels_path.exists():
        return index
    with open(parcels_path, "r", encoding="utf-8") as f:
        for line in f:
            try:
                obj = json.loads(line)
            except Exception:
                continue
            apn = obj.get("apn_norm") or obj.get("apn")
            addr = obj.get("site_address")
            if apn and addr:
                cleaned = clean_address(addr)
                if cleaned and cleaned not in index:
                    index[cleaned] = str(apn).strip().upper().replace("-", "")
    return index


# ---------------------------------------------------------------------------
# INPUT READERS
# ---------------------------------------------------------------------------

def iter_rows_csv(path: Path) -> Iterable[dict]:
    """Read CSV with utf-8-sig encoding (Excel compat)."""
    csv.field_size_limit(5_000_000)
    with open(path, "r", encoding="utf-8-sig", newline="") as f:
        reader = csv.DictReader(f)
        for row in reader:
            yield row


def iter_rows_xlsx(path: Path) -> Iterable[dict]:
    """Read XLSX via openpyxl."""
    try:
        from openpyxl import load_workbook
    except ImportError:
        print("[ERROR] openpyxl not installed. Run: pip install openpyxl", file=sys.stderr)
        sys.exit(1)
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows = ws.iter_rows(values_only=True)
    try:
        headers = [str(h or "").strip() for h in next(rows)]
    except StopIteration:
        return
    for row in rows:
        if not row or all(v is None for v in row):
            continue
        yield dict(zip(headers, row))


def detect_input(explicit: Optional[str] = None) -> Optional[Path]:
    """Find the recorder file. Respect explicit path, else try defaults."""
    if explicit:
        p = Path(explicit)
        if p.exists():
            return p
        print(f"[ERROR] Input not found: {explicit}", file=sys.stderr)
        sys.exit(2)
    for candidate in DEFAULT_INPUTS:
        if candidate.exists():
            return candidate
    return None


# ---------------------------------------------------------------------------
# COLUMN RESOLVER — handles casing/underscore variants
# ---------------------------------------------------------------------------

def get(row: dict, *keys: str) -> Optional[str]:
    """Fetch first non-empty value by a list of candidate keys (case-insensitive)."""
    lowered = {k.lower(): k for k in row.keys() if k}
    for key in keys:
        real_key = lowered.get(key.lower())
        if real_key is not None:
            v = row.get(real_key)
            if v is not None and str(v).strip():
                return str(v).strip()
    return None


# ---------------------------------------------------------------------------
# MAIN TRANSLATOR
# ---------------------------------------------------------------------------

def translate(input_path: Path, output_path: Path, dry_run: bool = False,
              parcels_path: Optional[Path] = None) -> dict:
    """Read recorder file, emit canonical JSONL. Returns summary stats."""
    stats = {
        "rows_read": 0,
        "emitted": 0,
        "matched_by_apn": 0,
        "matched_by_address": 0,
        "dropped_no_apn": 0,
        "dropped_unknown_doctype": 0,
        "by_signal": {},
    }

    # Build address index for fallback matching (only if parcels file exists)
    address_index: Dict[str, str] = {}
    if parcels_path is None:
        parcels_path = ROOT / "data" / "pima_parcels.jsonl"
    if parcels_path.exists():
        print(f"[INFO] Loading address index from {parcels_path.name}...")
        address_index = load_address_index(parcels_path)
        print(f"[INFO] Indexed {len(address_index):,} addresses for fallback matching")

    suffix = input_path.suffix.lower()
    if suffix == ".xlsx":
        rows = iter_rows_xlsx(input_path)
    elif suffix == ".csv":
        rows = iter_rows_csv(input_path)
    else:
        print(f"[ERROR] Unsupported file type: {suffix}", file=sys.stderr)
        sys.exit(2)

    out_handle = None if dry_run else open(output_path, "w", encoding="utf-8")

    try:
        for row in rows:
            stats["rows_read"] += 1

            doc_type_raw = get(row, "Document Type", "DocumentType", "doc_type", "Type")
            signal_type = map_doc_type(doc_type_raw) if doc_type_raw else None
            if not signal_type:
                stats["dropped_unknown_doctype"] += 1
                continue

            # Primary match: parcel ID column
            parcel_raw = get(row, "Parcel ID", "ParcelID", "APN", "parcel_id", "apn")
            apn_norm = norm_apn(parcel_raw) if parcel_raw else None
            match_source = "apn"

            # Fallback match: address lookup
            if not apn_norm and address_index:
                addr_raw = get(row, "Address", "address", "Property Address")
                cleaned = clean_address(addr_raw) if addr_raw else None
                if cleaned and cleaned in address_index:
                    apn_norm = address_index[cleaned]
                    match_source = "address"

            if not apn_norm:
                stats["dropped_no_apn"] += 1
                continue

            if match_source == "apn":
                stats["matched_by_apn"] += 1
            else:
                stats["matched_by_address"] += 1

            record = {
                "apn_norm": apn_norm,
                "apn_raw": parcel_raw,
                "county": "Pima",
                "signal_type": signal_type,
                "doc_type_raw": doc_type_raw,
                "record_date": get(row, "Recording Date", "RecordingDate", "record_date"),
                "doc_number": get(row, "Sequence Number", "SequenceNumber", "doc_number"),
                "grantors": get(row, "Grantors", "Grantor"),
                "grantees": get(row, "Grantees", "Grantee"),
                "address": get(row, "Address"),
                "legal_description": get(row, "Legal Description", "LegalDescription"),
                "source": "pima_recorder",
            }
            if out_handle:
                out_handle.write(json.dumps(record, ensure_ascii=False) + "\n")
            stats["emitted"] += 1
            stats["by_signal"][signal_type] = stats["by_signal"].get(signal_type, 0) + 1
    finally:
        if out_handle:
            out_handle.close()

    return stats


def main():
    parser = argparse.ArgumentParser(description="Translate Pima Recorder scraper output to JSONL signals")
    parser.add_argument("--input", "-i", help="Path to input CSV/XLSX (default: auto-detect)")
    parser.add_argument("--output", "-o", default=str(OUTPUT_JSONL), help=f"Output JSONL path (default: {OUTPUT_JSONL})")
    parser.add_argument("--parcels", "-p", help="Path to parcels JSONL for address fallback (default: data/pima_parcels.jsonl)")
    parser.add_argument("--dry-run", action="store_true", help="Preview only, don't write output file")
    args = parser.parse_args()

    input_path = detect_input(args.input)
    if not input_path:
        print(f"[ERROR] No recorder file found. Checked:", file=sys.stderr)
        for p in DEFAULT_INPUTS:
            print(f"        - {p}", file=sys.stderr)
        print(f"        Use --input PATH to specify a different location.", file=sys.stderr)
        sys.exit(2)

    print(f"[INFO] Reading:  {input_path}")
    if args.dry_run:
        print(f"[INFO] DRY RUN — no output file will be written")
    else:
        print(f"[INFO] Writing:  {args.output}")

    parcels_path = Path(args.parcels) if args.parcels else None
    stats = translate(input_path, Path(args.output), dry_run=args.dry_run, parcels_path=parcels_path)

    print()
    print(f"[DONE] Rows read:              {stats['rows_read']:>8,}")
    print(f"       Signals emitted:        {stats['emitted']:>8,}")
    print(f"         Matched by APN:       {stats['matched_by_apn']:>8,}")
    print(f"         Matched by address:   {stats['matched_by_address']:>8,}  (fallback)")
    print(f"       Dropped (no APN):       {stats['dropped_no_apn']:>8,}")
    print(f"       Dropped (unknown type): {stats['dropped_unknown_doctype']:>8,}")
    if stats["by_signal"]:
        print()
        print("       Signal breakdown:")
        for sig, n in sorted(stats["by_signal"].items(), key=lambda x: -x[1]):
            print(f"         {sig:<30} {n:>6,}")


if __name__ == "__main__":
    main()
