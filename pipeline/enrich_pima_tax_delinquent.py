#!/usr/bin/env python3
"""
Enrich Pima leads with tax-delinquent data from the HGL tax roll xlsx.

Input:  data/pima_tax_delinquent.xlsx  (HGL CRM Import sheet)
Output: data/tax_delinquent_pima.jsonl (consumed by build_leads.py)

Each output record has:
    apn_norm, county, signal_type="Tax Delinquent",
    amount_due, cert_amount, years_delinquent, tax_years, most_recent_year,
    foreclosure_imminence  (1-5 scale based on years_delinquent)

Usage:
    PYTHONPATH=. python pipeline/enrich_pima_tax_delinquent.py
"""
import json
import logging
from pathlib import Path
from openpyxl import load_workbook

BASE_DIR = Path(__file__).resolve().parent.parent
DATA_DIR = BASE_DIR / "data"

logging.basicConfig(level=logging.INFO, format="%(asctime)s [%(levelname)s] %(message)s",
                    datefmt="%H:%M:%S")
log = logging.getLogger("pima-tax")


def norm_apn(apn: str) -> str:
    """Pima APNs arrive as '219-21-627X' — normalize to the bare concatenation
       that matches the pima_parcels.jsonl apn_norm field."""
    if not apn:
        return ""
    return str(apn).replace("-", "").strip().upper()


def classify_foreclosure_imminence(years: int, amount: float) -> int:
    """
    Return a 1-5 scale of how likely this lead is to end up in foreclosure.
    Pima County typically initiates tax lien sales on properties 3+ years
    delinquent; 5+ years = almost certain auction candidate.
    """
    if years is None:
        return 1
    if years >= 10: return 5  # Deep distress, auction already likely
    if years >= 5:  return 4  # Very high risk
    if years >= 3:  return 3  # At the tax lien sale threshold
    if years >= 2:  return 2  # Early warning
    return 1


def main():
    src = DATA_DIR / "pima_tax_delinquent.xlsx"
    if not src.exists():
        log.error(f"Input file missing: {src}")
        log.error("Expected HGL tax delinquent xlsx at this path.")
        return

    log.info(f"Reading {src.name}...")
    wb = load_workbook(src, read_only=True)
    ws = wb["HGL CRM Import"]

    out_path = DATA_DIR / "tax_delinquent_pima.jsonl"
    count = 0
    high_imminence = 0
    with out_path.open("w") as f:
        for row in ws.iter_rows(min_row=2, values_only=True):
            # Columns: First, Last, Phone, PropAddr, City, State, Zip, Parcel, StateCode,
            #          CertAmount, TotalDue, YearsDel, TaxYears, MostRecent, ASR_URL
            apn_raw = row[7]
            if not apn_raw:
                continue
            apn_norm = norm_apn(apn_raw)
            cert_amt = row[9]
            total_due = row[10]
            years_del = row[11]
            tax_years = row[12]
            most_recent = row[13]
            asr_url = row[14]

            imminence = classify_foreclosure_imminence(years_del, total_due)
            if imminence >= 4:
                high_imminence += 1

            record = {
                "apn":           str(apn_raw),
                "apn_norm":      apn_norm,
                "county":        "Pima",
                "signal_type":   "Tax Delinquent",
                "amount":        total_due,
                "cert_amount":   cert_amt,
                "years_delinquent": years_del,
                "tax_years":     tax_years,
                "most_recent_year": most_recent,
                "foreclosure_imminence": imminence,
                "record_date":   f"{most_recent}-01-01" if most_recent else None,
                "doc_url":       asr_url,
                "source":        "HGL Pima tax delinquent xlsx",
            }
            f.write(json.dumps(record, default=str) + "\n")
            count += 1

    log.info(f"✓ wrote {count:,} Pima tax-delinquent signals → {out_path.name}")
    log.info(f"  high foreclosure imminence (4-5 scale): {high_imminence:,} parcels")


if __name__ == "__main__":
    main()
